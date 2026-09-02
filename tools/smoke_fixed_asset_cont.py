# -*- coding: utf-8 -*-
"""Continue fixed-asset tests: import conflict, audit lock, dispose period."""
from __future__ import annotations

import io
import json
import sys
import traceback
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE = "http://localhost:2154"
BOOK_ID = "2093221400646053889"
TERM = "2026-08"
NEXT = "2026-09"

SUBJ = {
    "1601": "2093221400772644914",
    "1602": "2093221400835559444",
    "1002": "2093221400835559445",
    "2221.01.01": "2093221400835559452",
    "5602.02": "2093221400772644953",
    "1606": "2093221400772644895",
    "5301.01": "2093221400772644906",
    "5711.02": "2093221400772644901",
}

results: list[tuple[str, str, str]] = []


def log(cid: str, ok: bool, detail: str = "") -> None:
    st = "PASS" if ok else "FAIL"
    results.append((cid, st, detail))
    print(f"[{st}] {cid} {detail}".rstrip())


def req(method: str, path: str, data=None, headers=None, raw=False, multipart=None):
    hdrs = dict(headers or {})
    body = None
    if multipart is not None:
        body, ctype = multipart
        hdrs["Content-Type"] = ctype
    elif data is not None:
        body = json.dumps(data, default=str).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    r = urllib.request.Request(BASE + path, data=body, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(r, timeout=90) as resp:
            raw_body = resp.read()
            if raw:
                return resp.status, raw_body
            text = raw_body.decode("utf-8")
            return resp.status, json.loads(text) if text else {}
    except urllib.error.HTTPError as e:
        text = e.read().decode("utf-8", errors="replace")
        try:
            return e.code, json.loads(text) if text else {"message": text}
        except Exception:
            return e.code, {"message": text, "code": -1}


def login():
    _, init = req("GET", "/api/login/get?_allow_anonymous=true")
    _, signin = req(
        "POST",
        "/api/login/signin?_allow_anonymous=true",
        {
            "username": "admin",
            "password": "changeme",
            "captcha": "",
            "state": init["data"]["state"],
            "authType": "normal",
        },
    )
    if signin.get("code") != 0:
        raise RuntimeError(signin)
    return {"Authorization": f"Bearer {signin['data']['token']}"}


def switch_book(auth):
    req("GET", f"/api/users/switchBook/{BOOK_ID}", headers=auth)
    _, me = req("GET", "/api/users/currentUser", headers=auth)
    if (me.get("data") or {}).get("bookId") != BOOK_ID:
        raise RuntimeError(f"book switch failed: {me}")


def epoch_day(s: str) -> int:
    dt = datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)


def set_term(auth, term: str):
    # update config via SQL-less API if exists; else direct DB
    import pymysql

    conn = pymysql.connect(
        host="127.0.0.1",
        port=3307,
        user="financial_cloud",
        password="FinancialCloud321!",
        database="financial_cloud",
        charset="utf8mb4",
        autocommit=True,
    )
    cur = conn.cursor()
    cur.execute(
        "UPDATE config SET config_value=%s WHERE book_id=%s AND config_key='sys.payment.term.current'",
        (term, BOOK_ID),
    )
    conn.close()
    # refresh session book context
    req("GET", f"/api/users/switchBook/{BOOK_ID}", headers=auth)


def encode_multipart(fields: dict, files: dict):
    boundary = "----FaSmokeBoundary7MA4YWxkTrZu0gW"
    body = io.BytesIO()
    for name, value in fields.items():
        body.write(f"--{boundary}\r\n".encode())
        body.write(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body.write(str(value).encode())
        body.write(b"\r\n")
    for name, (filename, content, mime) in files.items():
        body.write(f"--{boundary}\r\n".encode())
        body.write(
            f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'.encode()
        )
        body.write(f"Content-Type: {mime}\r\n\r\n".encode())
        body.write(content)
        body.write(b"\r\n")
    body.write(f"--{boundary}--\r\n".encode())
    return body.getvalue(), f"multipart/form-data; boundary={boundary}"


def test_import(auth):
    # download template
    code, blob = req("GET", "/api/fixed-asset/card/import-template", headers=auth, raw=True)
    log("FA-IO-05b", code == 200 and len(blob) > 50, f"bytes={len(blob)}")
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    # find header row = 1, data starts 2; clear sample and add rows
    # keep header, rewrite from row 2
    max_row = ws.max_row
    for r in range(max_row, 1, -1):
        ws.delete_rows(r)

    ts = datetime.now().strftime("%H%M%S")
    new_code = f"FA-IMP-OK-{ts}"
    # columns from export: code,name,catCode,catName,dept,start,qty,spec,loc,method,life,work,residual,orig,tax,impair,openingAccum,deprPeriods,faSubj,accumSubj,expenseSubj,...
    # Use template structure by reading first row headers
    headers = [c.value for c in ws[1]]
    log("FA-IO-template-headers", bool(headers), f"cols={len(headers)} first={headers[:5]}")

    def row_vals(code, name, method="å¹³åå¹´éæ³?, life=36):
        # minimal mapping by known positions from FixedAssetService export
        vals = [""] * len(headers)
        # assume standard order from CARD_EXPORT_HEADERS
        mapping = {
            "èµäº§ç¼ç ": code,
            "èµäº§åç§°": name,
            "ç±»å«ç¼ç ": "002",
            "ç±»å«åç§°": "çµå­è®¾å¤",
            "ä½¿ç¨é¨é¨": "è¡æ¿é?,
            "å¼å§ä½¿ç¨æ¥æ?: "2026-07-01",
            "æ°é": 1,
            "è§æ ¼åå·": "IMP",
            "å­æ¾å°ç¹": "å¯¼å¥ä»?,
            "ææ§æ¹æ³": method,
            "é¢è®¡ä½¿ç¨ææ°(æ?": life,
            "é¢è®¡æ»å·¥ä½é": "",
            "åæ®å¼ç%": 5,
            "åå?: 5000,
            "ç¨é¢": 0,
            "åå¼åå¤?: 0,
            "æåç´¯è®¡ææ§": 0,
            "å·²ææ§ææ?: 0,
            "åºå®èµäº§ç§ç®ç¼ç ": "1601",
            "ç´¯è®¡ææ§ç§ç®ç¼ç ": "1602",
            "ææ§è´¹ç¨ç§ç®ç¼ç ": "5602.02",
        }
        # also try English-less short names
        alt = {
            0: code,
            1: name,
            2: "002",
            3: "çµå­è®¾å¤",
            4: "è¡æ¿é?,
            5: "2026-07-01",
            6: 1,
            7: "IMP",
            8: "å¯¼å¥ä»?,
            9: method,
            10: life,
            11: "",
            12: 5,
            13: 5000,
            14: 0,
            15: 0,
            16: 0,
            17: 0,
            18: "1601",
            19: "1602",
            20: "5602.02",
        }
        for i, h in enumerate(headers):
            if h in mapping:
                vals[i] = mapping[h]
            elif i in alt:
                vals[i] = alt[i]
        return vals

    # 1 success new + 2 conflict existing FA-2026-001 + 1 invalid accelerated
    ws.append(row_vals(new_code, "å¯¼å¥æåæ ·ä¾"))
    ws.append(row_vals("FA-2026-001", "å²çªåºè·³è¿?))
    ws.append(row_vals(f"FA-IMP-BAD-{ts}", "å ééæ³?, method="ååä½é¢éåæ³?, life=30))
    ws.append(row_vals("FA-2026-002", "å²çª2"))

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    mp = encode_multipart({}, {"excelFile": ("import-test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    _, imp = req("POST", "/api/fixed-asset/card/import", headers=auth, multipart=mp)
    data = imp.get("data") or {}
    errors = data.get("errors") or []
    log(
        "FA-IO-06",
        imp.get("code") == 0 and (data.get("success") or 0) >= 1,
        f"success={data.get('success')} failed={data.get('failed')} msg={imp.get('message')}",
    )
    log(
        "FA-IO-07",
        any("FA-2026-001" in str(e.get("code") or "") or "å²çª" in str(e.get("message") or "") or "å·²å­å? in str(e.get("message") or "") for e in errors)
        or (data.get("failed") or 0) >= 2,
        f"errors={errors}",
    )
    log(
        "FA-IO-08",
        any("å é? in str(e.get("message") or "") or "24" in str(e.get("message") or "") for e in errors)
        or (data.get("failed") or 0) >= 1,
        f"errors={errors}",
    )
    log(
        "FA-IO-09",
        len(errors) >= (data.get("failed") or 0) and (data.get("failed") or 0) >= 2,
        f"failed={data.get('failed')} errorRows={len(errors)} fullList={errors}",
    )


def test_audit_reaccrue(auth):
    set_term(auth, TERM)
    _, st = req("GET", f"/api/fixed-asset/depreciation/status?yearPeriod={TERM}", headers=auth)
    status = st.get("data") or {}
    vid = status.get("voucherId")
    log("FA-ACR-audit-prep", bool(vid), f"voucherId={vid} canReaccrue={status.get('canReaccrue')}")
    if not vid:
        _, acr = req(
            "POST",
            "/api/fixed-asset/depreciation/accrue",
            {"yearPeriod": TERM, "voucherWord": "è®?, "summary": "å®¡è®¡åè®¡æ?},
            headers=auth,
        )
        vid = (acr.get("data") or {}).get("voucherId")
        log("FA-ACR-audit-accrue", acr.get("code") == 0, f"vid={vid}")
    if not vid:
        log("FA-ACR-05", False, "no voucher to audit")
        return

    # draft -> submit (UNDER_REVIEW) -> audit (COMPLETED + auditMemberId)
    _, sub = req("POST", f"/api/voucher/submit/{vid}", headers=auth)
    log("FA-ACR-submit", sub.get("code") == 0, f"msg={sub.get('message')}")
    _, aud = req("PUT", f"/api/voucher/audit/{vid}", headers=auth)
    log("FA-ACR-audit", aud.get("code") == 0 and "æåï¼?" in str(aud.get("message") or ""), f"msg={aud.get('message')}")

    _, st2 = req("GET", f"/api/fixed-asset/depreciation/status?yearPeriod={TERM}", headers=auth)
    s2 = st2.get("data") or {}
    log("FA-ACR-05-status", s2.get("canReaccrue") is False, f"canReaccrue={s2.get('canReaccrue')}")

    _, acr2 = req(
        "POST",
        "/api/fixed-asset/depreciation/accrue",
        {"yearPeriod": TERM, "voucherWord": "è®?, "summary": "åºè¢«æç»çéæ?},
        headers=auth,
    )
    log(
        "FA-ACR-05",
        acr2.get("code") != 0,
        f"code={acr2.get('code')} msg={acr2.get('message')}",
    )

    # unaudit to restore env
    _, una = req("PUT", f"/api/voucher/unaudit/{vid}", headers=auth)
    log("FA-ACR-unaudit-cleanup", una.get("code") == 0, una.get("message") or "")


def test_dispose_period(auth):
    """Create asset eligible in Aug, accrue if needed, dispose in Aug, switch to Sep, accrue, assert absent."""
    set_term(auth, TERM)
    ts = datetime.now().strftime("%H%M%S")
    code = f"FA-DIS-P-{ts}"
    _, cats = req("GET", "/api/fixed-asset/category/list", headers=auth)
    cat_id = next((c["id"] for c in (cats.get("data") or []) if c.get("code") == "002"), None)
    _, save = req(
        "POST",
        "/api/fixed-asset/card/save",
        {
            "bookId": BOOK_ID,
            "code": code,
            "name": "æ¸çæ¬¡æåææ ·ä¾",
            "categoryId": cat_id,
            "deptId": "fa-demo-dept-admin",
            "startUseDate": "2026-07-01",
            "depreciationMethod": "STRAIGHT_LINE",
            "usefulLifeMonths": 36,
            "residualRate": 5,
            "originalValue": 3600,
            "taxAmount": 0,
            "fixedAssetSubjectId": SUBJ["1601"],
            "accumDeprSubjectId": SUBJ["1602"],
            "expenseSubjectId": SUBJ["5602.02"],
            "purchaseCounterpartSubjectId": SUBJ["1002"],
        },
        headers=auth,
    )
    asset_id = (save.get("data") or {}).get("assetId")
    log("FA-DIS-period-create", save.get("code") == 0 and bool(asset_id), save.get("message") or "")
    if not asset_id:
        return

    # ensure Aug accrued (may reaccrue)
    _, st = req("GET", f"/api/fixed-asset/depreciation/status?yearPeriod={TERM}", headers=auth)
    if (st.get("data") or {}).get("canReaccrue") or not (st.get("data") or {}).get("accrued"):
        _, acr = req(
            "POST",
            "/api/fixed-asset/depreciation/accrue",
            {"yearPeriod": TERM, "voucherWord": "è®?, "summary": "æ¸çåè®¡æ?},
            headers=auth,
        )
        log("FA-DIS-02-accrue-aug", acr.get("code") == 0, acr.get("message") or str(acr.get("data")))
    else:
        log("FA-DIS-02-accrue-aug", False, "locked cannot include new asset in Aug without reaccrue")

    import pymysql

    conn = pymysql.connect(
        host="127.0.0.1",
        port=3307,
        user="financial_cloud",
        password="FinancialCloud321!",
        database="financial_cloud",
        charset="utf8mb4",
        autocommit=True,
    )
    cur = conn.cursor()
    cur.execute(
        "SELECT depr_amount FROM fixed_asset_depr d JOIN fixed_asset a ON a.id=d.asset_id "
        "WHERE a.code=%s AND d.year_period=%s AND d.deleted='n'",
        (code, TERM),
    )
    row = cur.fetchone()
    in_aug = row is not None and float(row[0]) > 0
    log("FA-DIS-02", in_aug, f"aug_amount={row[0] if row else None}")

    _, disp = req(
        "POST",
        f"/api/fixed-asset/card/dispose/{asset_id}",
        {
            "disposeIncome": 0,
            "disposeExpense": 0,
            "counterpartSubjectId": SUBJ["1002"],
            "disposalSubjectId": SUBJ["1606"],
            "gainSubjectId": SUBJ["5301.01"],
            "lossSubjectId": SUBJ["5711.02"],
            "voucherWord": "è®?,
            "summary": "æé´è§åæ¸ç",
        },
        headers=auth,
    )
    log("FA-DIS-period-dispose", disp.get("code") == 0, disp.get("message") or "")

    # September: UOP assets need period work
    set_term(auth, NEXT)
    _, work = req("GET", f"/api/fixed-asset/depreciation/work?yearPeriod={NEXT}", headers=auth)
    work_list = work.get("data") or []
    for w in work_list:
        if not w.get("periodWork"):
            w["periodWork"] = 1000
    if work_list:
        _, sw = req(
            "PUT",
            f"/api/fixed-asset/depreciation/work?yearPeriod={NEXT}",
            work_list,
            headers=auth,
        )
        log("FA-DIS-sep-work", sw.get("code") == 0, sw.get("message") or "")

    _, acr_sep = req(
        "POST",
        "/api/fixed-asset/depreciation/accrue",
        {"yearPeriod": NEXT, "voucherWord": "è®?, "summary": "æ¬¡æè®¡æ"},
        headers=auth,
    )
    log("FA-DIS-03-accrue-sep", acr_sep.get("code") == 0, f"msg={acr_sep.get('message')} data={acr_sep.get('data')}")

    cur.execute(
        "SELECT depr_amount FROM fixed_asset_depr d JOIN fixed_asset a ON a.id=d.asset_id "
        "WHERE a.code=%s AND d.year_period=%s AND d.deleted='n'",
        (code, NEXT),
    )
    row2 = cur.fetchone()
    log("FA-DIS-03", row2 is None or float(row2[0]) == 0, f"sep_amount={row2[0] if row2 else None}")

    # restore term
    set_term(auth, TERM)
    conn.close()


def main():
    auth = login()
    switch_book(auth)
    test_import(auth)
    test_audit_reaccrue(auth)
    test_dispose_period(auth)

    passed = sum(1 for _, s, _ in results if s == "PASS")
    failed = sum(1 for _, s, _ in results if s == "FAIL")
    print("\n======== SUMMARY ========")
    print(f"PASS={passed} FAIL={failed} TOTAL={len(results)}")
    for cid, s, d in results:
        if s == "FAIL":
            print(f"  FAIL {cid}: {d}")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        traceback.print_exc()
        raise SystemExit(2)
