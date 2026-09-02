#!/usr/bin/env python3
"""Import standard accounting subjects from official xlsx templates into SQL seed files."""
from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
OUT_SQL = ROOT / "sql" / "seed" / "data" / "standard_subjects.sql"
OUT_REPORT = ROOT / "docs" / "subject-import-compatibility.md"

STANDARDS = [
    ("1", "小企业会计准则", DOCS / "小企业会计准则科目.xlsx"),
    ("2", "企业会计制度", DOCS / "企业会计制度科目.xlsx"),
]

CATEGORY_MAP = {
    "流动资产": 1,
    "长期资产": 1,
    "流动负债": 2,
    "长期负债": 2,
    "所有者权益": 4,
    "成本": 5,
    "营业成本及税金": 5,
    "营业收入": 6,
    "期间费用": 6,
    "其他收益": 6,
    "其他损失": 6,
    "所得税": 6,
    "以前年度损益调整": 6,
}


def make_id(standard_id: str, code: str) -> str:
    digest = hashlib.sha256(f"{standard_id}:{code}".encode("utf-8")).hexdigest()
    return str(int(digest[:15], 16))


def parent_code(code: str) -> str | None:
    if "." in code:
        parent = code.rsplit(".", 1)[0]
        return parent if parent else None
    length = len(code)
    if length == 6:
        return code[:4]
    if length == 8:
        return code[:6]
    if length == 12:
        return code[:8]
    return None


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def yn_flag(cell) -> int:
    return 1 if cell and str(cell).strip() == "是" else 0


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else code
        category_text = str(row[3]).strip() if row[3] else "流动资产"
        category = CATEGORY_MAP.get(category_text, 1)
        direction = "1" if row[4] and str(row[4]).strip() == "借" else "2"
        display_name = str(row[6]).strip() if row[6] else name
        pinyin = str(row[2]).strip() if row[2] else None
        is_cash = yn_flag(row[11]) or yn_flag(row[12])
        rows.append(
            {
                "code": code,
                "name": name,
                "display_name": display_name,
                "pinyin_code": pinyin,
                "category": category,
                "direction": direction,
                "is_cash": is_cash,
            }
        )
    wb.close()
    return rows


def subject_level(code: str, has_parent: bool) -> int:
    if not has_parent:
        return 1
    if "." in code:
        return code.count(".") + 1
    return {6: 2, 8: 3, 12: 4}.get(len(code), 1)


def build_id_path(code: str, id_by_code: dict[str, str]) -> str:
    chain: list[str] = []
    current: str | None = code
    while current and current in id_by_code:
        chain.append(current)
        current = parent_code(current)
        if current is not None and current not in id_by_code:
            break
    chain.reverse()
    return "/" + "/".join(id_by_code[item] for item in chain)


def build_subjects(standard_id: str, rows: list[dict]) -> list[dict]:
    by_code = {row["code"]: row for row in rows}
    ids: dict[str, str] = {row["code"]: make_id(standard_id, row["code"]) for row in rows}
    subjects: list[dict] = []
    for row in rows:
        code = row["code"]
        parent = parent_code(code)
        has_parent = bool(parent and parent in by_code)
        subjects.append(
            {
                **row,
                "id": ids[code],
                "standard_id": standard_id,
                "parent_code": parent if has_parent else None,
                "level": subject_level(code, has_parent),
                "status": 1,
                "system_default": 1,
                "auxiliary": "[]",
            }
        )
    id_by_code = {s["code"]: s["id"] for s in subjects}
    for subject in subjects:
        if subject["parent_code"]:
            subject["parent_id"] = id_by_code[subject["parent_code"]]
            subject["id_path"] = build_id_path(subject["code"], id_by_code)
        else:
            subject["parent_id"] = None
            subject["id_path"] = f"/{subject['id']}"
    subjects.sort(key=lambda s: s["code"])
    return subjects


def compare_standards(all_subjects: dict[str, list[dict]]) -> dict:
    small = {s["code"]: s["name"] for s in all_subjects["1"]}
    ent = {s["code"]: s["name"] for s in all_subjects["2"]}
    shared = sorted(set(small) & set(ent))
    return {
        "small_count": len(small),
        "enterprise_count": len(ent),
        "only_small": sorted(set(small) - set(ent))[:30],
        "only_enterprise": sorted(set(ent) - set(small))[:30],
        "renamed": [(c, small[c], ent[c]) for c in shared if small[c] != ent[c]][:30],
        "code_format_examples": {
            "dotted_small": [c for c in small if "." in c][:8],
            "legacy_style": [c for c in small if "." not in c and len(c) > 4][:8],
        },
    }


def write_sql(all_subjects: dict[str, list[dict]]) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "-- Generated by tools/import_standard_subjects.py",
        f"-- Generated at: {now}",
        "SET NAMES utf8mb4;",
        "",
        "DELETE FROM standard_subject_cash_flow;",
        "DELETE FROM standard_subject WHERE standard_id IN ('1','2');",
        "UPDATE standard SET name = '小企业会计准则' WHERE id = '1';",
        "UPDATE standard SET name = '企业会计制度' WHERE id = '2';",
        "",
    ]
    for standard_id, _name, _path in STANDARDS:
        lines.append(f"-- standard_id={standard_id} {_name}")
        for s in all_subjects[standard_id]:
            cols = (
                s["id"],
                s["category"],
                s["code"],
                s["name"],
                s["display_name"],
                s.get("pinyin_code"),
                s.get("pinyin_code"),
                s["direction"],
                s["status"],
                s["parent_id"],
                s["id_path"],
                s["level"],
                s["system_default"],
                s["is_cash"],
                None,
                None,
                s["auxiliary"],
                None,
                None,
                None,
                s["standard_id"],
                "1",
                now,
                "1",
                now,
                "n",
            )
            values = ", ".join(
                [
                    sql_str(cols[0]),
                    str(cols[1]),
                    sql_str(cols[2]),
                    sql_str(cols[3]),
                    sql_str(cols[4]),
                    sql_str(cols[5]),
                    sql_str(cols[6]),
                    sql_str(cols[7]),
                    str(cols[8]),
                    sql_str(cols[9]),
                    sql_str(cols[10]),
                    str(cols[11]),
                    str(cols[12]),
                    str(cols[13]),
                    "NULL",
                    "NULL",
                    sql_str(cols[16]),
                    "NULL",
                    "NULL",
                    "NULL",
                    sql_str(cols[20]),
                    sql_str(cols[21]),
                    sql_str(cols[22]),
                    sql_str(cols[23]),
                    sql_str(cols[24]),
                    sql_str(cols[25]),
                ]
            )
            lines.append(f"INSERT INTO standard_subject VALUES ({values});")
        lines.append("")
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")


def write_report(compare: dict) -> None:
    lines = [
        "# 会计科目导入兼容性说明",
        "",
        "> 由 `tools/import_standard_subjects.py` 自动生成",
        "",
        "## 数据量",
        "",
        f"- 小企业会计准则（standard_id=1）：**{compare['small_count']}** 条",
        f"- 企业会计制度（standard_id=2）：**{compare['enterprise_count']}** 条",
        "",
        "## 编码格式变化",
        "",
        "- 新模板大量使用 **点分编码**（如 `1012.01`），旧系统多为 **定长拼接**（如 `101201`）。",
        "- 后端已支持 `[\\d\\-.]+` 校验；`reorgSubjectName` 需同时支持点分与旧规则。",
        "",
        "## 同码不同名（前 30 条）",
        "",
        "| 编码 | 小企业名称 | 企业会计制度名称 |",
        "|------|------------|------------------|",
    ]
    for code, n1, n2 in compare["renamed"]:
        lines.append(f"| {code} | {n1} | {n2} |")
    lines.extend(
        [
            "",
            "## 仅小企业准则存在的编码（示例）",
            "",
            ", ".join(compare["only_small"]) or "无",
            "",
            "## 仅企业会计制度存在的编码（示例）",
            "",
            ", ".join(compare["only_enterprise"]) or "无",
            "",
            "## 旧编码 → 新编码（业务硬编码对照）",
            "",
            "| 旧编码（定长子码） | 新编码（点分） | 说明 |",
            "|------------------|--------------|------|",
            "| 101201 | 1012.01 | 其他货币资金明细（旧码因准则不同可能对应 1012.05 外埠存款） |",
            "| 122102 | 1221 | 其他应收款（新模板无 1221.02 子目） |",
            "| 221101 | 2211.01 | 应付职工薪酬_职工工资 |",
            "| 221103 | 2211.04 | 应付职工薪酬_社会保险费 |",
            "| 222114 | 2221.14 | 应交税费_个人所得税 |",
            "| 224101 | 2241 | 其他应付款（新模板无 2241.01 子目） |",
            "| 660222 | 4002 | 劳务费场景映射至劳务成本 |",
            "",
            "后端 `SubjectCodeCompat` 会在凭证模板查找时依次尝试新码与旧码。",
            "",
            "## 影响范围",
            "",
            "1. **standard_subject**：本脚本会全量替换 standard_id 1/2。",
            "2. **book_subject**：已有账套科目不会自动更新；需重建账套或手工同步。",
            "3. **凭证/余额**：引用旧 subject_id 的历史数据需迁移或清空。",
            "4. **config 默认科目**（如 1001/1002/1122）：编码未变则仍兼容，变更编码需同步 config 表。",
            "5. **standard_subject_cash_flow**：已清空，需按新科目重新配置。",
            "",
            "## 执行",
            "",
            "```bash",
            "python tools/import_standard_subjects.py",
            "mysql ... < sql/seed/data/standard_subjects.sql",
            "```",
            "",
        ]
    )
    OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    all_subjects: dict[str, list[dict]] = {}
    for standard_id, name, path in STANDARDS:
        if not path.exists():
            print(f"Missing file: {path}", file=sys.stderr)
            return 1
        rows = load_rows(path)
        subjects = build_subjects(standard_id, rows)
        all_subjects[standard_id] = subjects
        print(f"{name}: {len(subjects)} subjects")
    compare = compare_standards(all_subjects)
    write_sql(all_subjects)
    write_report(compare)
    print(f"Wrote {OUT_SQL}")
    print(f"Wrote {OUT_REPORT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
